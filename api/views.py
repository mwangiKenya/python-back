from django.http import JsonResponse, HttpResponse
from .models import (
    read_users, readings, Admin, Billings, Logs, Users, history,
    ReadingHistory, PaymentHistory, BillingHistory, AuditTrail, 
    BillingCycleHistory, CustomerPaymentSummary
)
from django.views.decorators.csrf import csrf_exempt
import json
import secrets
from django.db import transaction
from datetime import date
from decimal import Decimal
from django.db.models import Sum, Avg, Count, Q
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .sms import send_sms
import calendar
from datetime import datetime, date, timedelta
import pandas as pd
from django.db import transaction
from django.http import JsonResponse
from datetime import date, datetime
from .models import read_users, readings, Billings, Logs
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from django.conf import settings
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta
from django.utils import timezone
import requests
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from io import BytesIO

#======================================================================================
# NEW HELPER FUNCTIONS FOR HISTORY TRACKING
#======================================================================================

def create_reading_history(reading, recorded_by="system", role="system"):
    """Create a historical record of a reading"""
    try:
        ReadingHistory.objects.create(
            reading_id=reading.id,
            user_id=reading.user_id,
            name=reading.name,
            phone=reading.phone,
            metre_num=reading.metre_num,
            grp=reading.grp,
            parent=reading.parent,
            prev_user=reading.prev_user or 0,
            prev_sup=reading.prev_sup or 0,
            cur_user=reading.cur_user or 0,
            cur_sup=reading.cur_sup or 0,
            mid_user=reading.mid_user,
            mid_sup=reading.mid_sup,
            units_used=reading.units_used or 0,
            rate=reading.rate or 0,
            reading_date=reading.cur_date or date.today(),
            prev_date=reading.prev_date,
            cycle_month=reading.cur_date.strftime("%Y-%m") if reading.cur_date else None,
            recorded_by=recorded_by,
            role=role,
            version=getattr(reading, 'version', 1)
        )
    except Exception as e:
        print(f"Error creating reading history: {e}")

def create_payment_history(billing, amount, previous_balance, 
                          payment_method='CASH', recorded_by="system", 
                          role="system", notes=None):
    """Create a payment history record with corrected math"""
    try:
        receipt_number = f"RCP-{datetime.now().strftime('%Y%m%d')}-{billing.id}-{secrets.token_hex(4).upper()}"
        
        # Calculate current balance correctly
        # current_balance = previous_balance - amount_paid
        current_balance = previous_balance - amount
        
        # Determine payment status
        if amount >= previous_balance:
            status = 'COMPLETED'
        else:
            status = 'PARTIAL'
        
        PaymentHistory.objects.create(
            billing_id=billing.id,
            user_id=billing.user_id,
            name=billing.name,
            phone=billing.phone,
            grp=billing.grp,
            parent=billing.parent,
            amount_paid=amount,
            previous_balance=previous_balance,
            current_balance=current_balance,
            bill_amount=billing.bill,
            payment_method=payment_method,
            status=status,
            receipt_number=receipt_number,
            notes=notes,
            recorded_by=recorded_by,
            role=role
        )
        return receipt_number
    except Exception as e:
        print(f"Error creating payment history: {e}")
        return None

def create_billing_history(billing, cycle_month, generated_by="system", role="system"):
    """Create a billing history record"""
    try:
        BillingHistory.objects.create(
            billing_id=billing.id,
            user_id=billing.user_id,
            name=billing.name,
            phone=billing.phone,
            metre_num=billing.sms_name,
            grp=billing.grp,
            parent=billing.parent,
            units_used=billing.units_used,
            rate=billing.rate,
            current_bill=billing.bill,
            previous_balance=billing.b_cd,
            total_due=billing.bill + billing.b_cd,
            amount_paid=billing.paid,
            remaining_balance=billing.bal,
            prev_reading=billing.prev_user,
            current_reading=billing.cur_user,
            cycle_month=cycle_month,
            billing_date=billing.billed_on or date.today(),
            due_date=(billing.billed_on or date.today()) + timedelta(days=30),
            status=billing.status,
            generated_by=generated_by,
            role=role
        )
    except Exception as e:
        print(f"Error creating billing history: {e}")

def update_customer_summary(user_id):
    """Update the customer payment summary"""
    try:
        billing = Billings.objects.filter(user_id=user_id).first()
        if billing:
            summary, created = CustomerPaymentSummary.objects.get_or_create(
                user_id=user_id,
                defaults={
                    'name': billing.name,
                    'phone': billing.phone,
                    'metre_num': billing.sms_name,
                    'grp': billing.grp,
                    'parent': billing.parent,
                }
            )
            # Update summary with aggregated data
            total_billed = Billings.objects.filter(user_id=user_id).aggregate(
                Sum('bill')
            )['bill__sum'] or 0
            
            total_paid = Billings.objects.filter(user_id=user_id).aggregate(
                Sum('paid')
            )['paid__sum'] or 0
            
            summary.total_billed = total_billed
            summary.total_paid = total_paid
            summary.current_balance = billing.bal
            
            # Update last payment info
            last_payment = PaymentHistory.objects.filter(
                user_id=user_id,
                status='COMPLETED'
            ).order_by('-timestamp').first()
            
            if last_payment:
                summary.last_payment_date = last_payment.payment_date
                summary.last_payment_amount = last_payment.amount_paid
                summary.payment_count = PaymentHistory.objects.filter(
                    user_id=user_id,
                    status='COMPLETED'
                ).count()
            
            # Update status
            if billing.bal <= 0:
                summary.payment_status = 'PAID'
            elif billing.bal > 0 and billing.bal < 1000:
                summary.payment_status = 'CURRENT'
            elif billing.bal >= 1000 and billing.bal < 5000:
                summary.payment_status = 'OVERDUE'
            else:
                summary.payment_status = 'DELINQUENT'
            
            summary.save()
    except Exception as e:
        print(f"Error updating customer summary: {e}")

def create_audit_trail(username, role, action, table_name=None, record_id=None,
                       field_changed=None, old_value=None, new_value=None,
                       description=None, request=None):
    """Create an audit trail entry"""
    try:
        ip_address = None
        user_agent = None
        session_id = None
        
        if request:
            ip_address = request.META.get('REMOTE_ADDR')
            user_agent = request.META.get('HTTP_USER_AGENT')
            session_id = request.session.session_key
        
        AuditTrail.objects.create(
            username=username,
            role=role,
            action=action,
            table_name=table_name,
            record_id=record_id,
            field_changed=field_changed,
            old_value=old_value,
            new_value=new_value,
            description=description or f"{action} performed by {username}",
            ip_address=ip_address,
            user_agent=user_agent,
            session_id=session_id
        )
    except Exception as e:
        print(f"Error creating audit trail: {e}")

def update_reading_field(reading, field_name, new_value, username="system", role="system"):
    """
    Safely updates a field AND logs history automatically
    """
    old_value = (
        readings.objects
        .filter(id=reading.id)
        .values_list(field_name, flat=True)
        .first()
    )
    
    # only log if something actually changes
    if old_value != new_value:
        history.objects.create(
            name=reading.name,
            field=field_name,
            old_val=old_value if old_value is not None else 0,
            new_val=new_value if new_value is not None else 0
        )
        create_log(
            username=username,
            role=role,
            action="UPDATE",
            table="readings",
            record_id=reading.id,
            field_changed=field_name,
            old_val=old_value,
            new_val=new_value,
            description=f"{field_name} updated for {reading.name}"
        )
        create_audit_trail(
            username=username,
            role=role,
            action="UPDATE",
            table_name="readings",
            record_id=reading.id,
            field_changed=field_name,
            old_value=old_value,
            new_value=new_value,
            description=f"{field_name} updated for {reading.name}"
        )
    
    setattr(reading, field_name, new_value)

#======================================================================================
# EXISTING GLOBALS AND HELPERS
#======================================================================================

CYCLE_SCHEDULER = {
    "end_time": None
}
BILLING_STATE = {
    "start_month": None,
}
CYCLE_CONFIG = {
    "start_date": None,
    "delay_days": 30,
}
BILLING_CYCLE = {
    "start_month": 5,
    "start_year": 2026,
    "shift_days": 0
}
LAST_STATE_SNAPSHOT = None

def last_day_of_month(year, month):
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day)

def create_log(username, role, action, table, record_id, description,
               field_changed=None, old_val=None, new_val=None):
    Logs.objects.create(
        username=username,
        role=role,
        action=action,
        table_name=table,
        record_id=record_id,
        field_changed=field_changed,
        old_val=str(old_val) if old_val is not None else None,
        new_val=str(new_val) if new_val is not None else None,
        description=description
    )

def create_hist(name, field, old_val, new_val):
    history.objects.create(
        name=name,
        field=field,
        old_val=old_val,
        new_val=new_val
    )

def snapshot_readings():
    global LAST_STATE_SNAPSHOT
    LAST_STATE_SNAPSHOT = list(readings.objects.values())

#======================================================================================
# CYCLE MANAGEMENT ENDPOINTS
#======================================================================================

@csrf_exempt
def set_cycle_duration(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        days = int(data.get("days", 0))
        hours = int(data.get("hours", 0))
        minutes = int(data.get("minutes", 0))
        seconds = int(data.get("seconds", 0))
        delta = timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds)
        end_time = timezone.now() + delta
        CYCLE_SCHEDULER["end_time"] = end_time
        return JsonResponse({
            "message": "Cycle timer started",
            "end_time": end_time.isoformat()
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def cycle_timer_status(request):
    now = timezone.now()
    if not CYCLE_SCHEDULER["end_time"]:
        return JsonResponse({
            "running": False,
            "days": 0,
            "hours": 0,
            "minutes": 0,
            "seconds": 0
        })
    diff = CYCLE_SCHEDULER["end_time"] - now
    if diff.total_seconds() <= 0:
        return JsonResponse({
            "running": False,
            "expired": True
        })
    return JsonResponse({
        "running": True,
        "days": diff.days,
        "hours": diff.seconds // 3600,
        "minutes": (diff.seconds % 3600) // 60,
        "seconds": diff.seconds % 60
    })

@csrf_exempt
def auto_shift_if_due(request):
    now = timezone.now()
    if not CYCLE_SCHEDULER["end_time"]:
        return JsonResponse({"message": "No cycle running"})
    if now < CYCLE_SCHEDULER["end_time"]:
        return JsonResponse({"message": "Not yet time"})
    
    CYCLE_SCHEDULER["end_time"] = None
    
    with transaction.atomic():
        next_month = now.month + 1
        next_year = now.year
        if next_month > 12:
            next_month = 1
            next_year += 1
        next_last_day = calendar.monthrange(next_year, next_month)[1]
        next_cycle_date = date(next_year, next_month, next_last_day)
        
        for r in readings.objects.all():
            # Create history before shifting
            create_reading_history(r, "system", "system")
            
            r.prev_user = r.cur_user if r.cur_user is not None else r.prev_user
            r.prev_sup = r.cur_sup if r.cur_sup is not None else r.prev_sup
            r.prev_date = r.cur_date or r.prev_date
            r.cur_date = next_cycle_date
            r.cur_user = None
            r.cur_sup = None
            r.save()
    
    return JsonResponse({
        "message": "Auto shift completed",
        "next_cycle_date": str(next_cycle_date)
    })

@csrf_exempt
def start_billing_month(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        start_month = data.get("start_month")
        if not start_month:
            return JsonResponse({"error": "start_month required"}, status=400)
        
        year, month = map(int, start_month.split("-"))
        last_day = calendar.monthrange(year, month)[1]
        prev_date = date(year, month, last_day)
        
        next_month = month + 1
        next_year = year
        if next_month > 12:
            next_month = 1
            next_year += 1
        next_last_day = calendar.monthrange(next_year, next_month)[1]
        cur_date = date(next_year, next_month, next_last_day)
        
        with transaction.atomic():
            readings.objects.all().update(
                prev_date=prev_date,
                cur_date=cur_date
            )
            
            # Create billing cycle record
            BillingCycleHistory.objects.create(
                cycle_month=start_month,
                start_date=prev_date,
                end_date=cur_date,
                next_cycle_date=cur_date + timedelta(days=30),
                status='IN_PROGRESS',
                started_by=data.get("username", "system")
            )
        
        BILLING_STATE["start_month"] = start_month
        return JsonResponse({
            "message": "Billing month started",
            "prev_date": str(prev_date),
            "cur_date": str(cur_date)
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# FETCH ENDPOINTS
#======================================================================================

def water_users(request):
    users = read_users.objects.all()
    data = []
    for u in users:
        data.append({
            'id': u.id,
            'fname': u.fname,
            'phone': u.phone,
            'metre_num': u.metre_num,
            'zone': u.zone,
            'rate': u.rate,
            'created_on': u.created_on.strftime('%Y-%m-%d') if u.created_on else None,
            'grp': u.grp,
            'parent': u.parent
        })
    return JsonResponse(data, safe=False)

def hist_data(request):
    name = request.GET.get("name")
    field = request.GET.get("field")
    hist = history.objects.all()
    if name:
        hist = hist.filter(name__icontains=name)
    if field:
        hist = hist.filter(field__icontains=field)
    data = []
    for h in hist:
        data.append({
            'id': h.id,
            'name': h.name,
            'field': h.field,
            'old_val': h.old_val,
            'new_val': h.new_val,
            'changes_on': h.changed_on.strftime('%Y-%m-%d') if h.changed_on else None
        })
    return JsonResponse(data, safe=False)

def bill(request):
    bills = Billings.objects.all()
    data = []
    for b in bills:
        data.append({
            'id': b.id,
            'user_id': b.user_id,
            'name': b.name,
            'phone': b.phone,
            'units_used': b.units_used,
            'rate': b.rate,
            'bill': b.bill,
            'paid': b.paid,
            'bal': b.bal,
            'status': b.status,
            'b_cd': b.b_cd,
            'prev_user': b.prev_user,
            'cur_user': b.cur_user,
            'sms_name': b.sms_name,
            'grp': b.grp,
            'parent': b.parent
        })
    return JsonResponse(data, safe=False)

def logs(request):
    log = Logs.objects.all().order_by('-changed_at')
    data = []
    for l in log:
        data.append({
            'id': l.id,
            'username': l.username,
            'role': l.role,
            'action': l.action,
            'table_name': l.table_name,
            'record_id': l.record_id,
            'field_changed': l.field_changed,
            'old_val': l.old_val,
            'new_val': l.new_val,
            'description': l.description,
            'changed_at': l.changed_at.strftime('%Y-%m-%d %H:%M:%S') if l.changed_at else None
        })
    return JsonResponse(data, safe=False)

def read_data(request):
    read = readings.objects.all()
    data = []
    for r in read:
        data.append({
            'id': r.id,
            'user_id': r.user_id,
            'name': r.name,
            'phone': r.phone,
            'metre_num': r.metre_num,
            'prev_user': r.prev_user,
            'prev_sup': r.prev_sup,
            'prev_date': r.prev_date.strftime('%Y-%m-%d') if r.prev_date else None,
            'cur_user': r.cur_user,
            'cur_sup': r.cur_sup,
            'cur_date': r.cur_date.strftime('%Y-%m-%d') if r.cur_date else None,
            'rate': r.rate,
            'mid_user': r.mid_user,
            'mid_sup': r.mid_sup,
            'grp': r.grp,
            'parent': r.parent
        })
    return JsonResponse(data, safe=False)

#======================================================================================
# NEW HISTORY FETCH ENDPOINTS
#======================================================================================

@api_view(['GET'])
def get_reading_history(request):
    """Get reading history with filters"""
    user_id = request.GET.get('user_id')
    cycle_month = request.GET.get('cycle_month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    history_qs = ReadingHistory.objects.all()
    
    if user_id:
        history_qs = history_qs.filter(user_id=user_id)
    if cycle_month:
        history_qs = history_qs.filter(cycle_month=cycle_month)
    if start_date:
        history_qs = history_qs.filter(timestamp__date__gte=start_date)
    if end_date:
        history_qs = history_qs.filter(timestamp__date__lte=end_date)
    
    data = list(history_qs.values(
        'id', 'name', 'phone', 'prev_user', 'cur_user', 
        'units_used', 'cycle_month', 'timestamp', 'recorded_by'
    ))
    return Response(data)

@api_view(['GET'])
def get_payment_history(request):
    """Get payment history with filters"""
    user_id = request.GET.get('user_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    history_qs = PaymentHistory.objects.all()
    
    if user_id:
        history_qs = history_qs.filter(user_id=user_id)
    if start_date:
        history_qs = history_qs.filter(payment_date__gte=start_date)
    if end_date:
        history_qs = history_qs.filter(payment_date__lte=end_date)
    
    data = list(history_qs.values(
        'id', 'name', 'phone', 'amount_paid', 'previous_balance',
        'current_balance', 'payment_method', 'receipt_number',
        'payment_date', 'recorded_by'
    ))
    return Response(data)

@api_view(['GET'])
def get_billing_history(request):
    """Get billing history with filters"""
    user_id = request.GET.get('user_id')
    cycle_month = request.GET.get('cycle_month')
    status = request.GET.get('status')
    
    history_qs = BillingHistory.objects.all()
    
    if user_id:
        history_qs = history_qs.filter(user_id=user_id)
    if cycle_month:
        history_qs = history_qs.filter(cycle_month=cycle_month)
    if status:
        history_qs = history_qs.filter(status=status)
    
    data = list(history_qs.values(
        'id', 'name', 'phone', 'units_used', 'current_bill',
        'total_due', 'amount_paid', 'remaining_balance',
        'cycle_month', 'status', 'billing_date', 'due_date'
    ))
    return Response(data)

@api_view(['GET'])
def get_customer_history(request, user_id):
    """Get complete history for a customer"""
    try:
        reading_history = ReadingHistory.objects.filter(
            user_id=user_id
        ).values('timestamp', 'cur_user', 'prev_user', 'units_used', 'cycle_month', 'recorded_by')
        
        payment_history = PaymentHistory.objects.filter(
            user_id=user_id
        ).values('timestamp', 'amount_paid', 'previous_balance', 
                'current_balance', 'payment_method', 'receipt_number', 'payment_date')
        
        billing_history = BillingHistory.objects.filter(
            user_id=user_id
        ).values('cycle_month', 'current_bill', 'total_due', 'amount_paid', 
                'remaining_balance', 'status')
        
        return Response({
            'user_id': user_id,
            'reading_history': reading_history,
            'payment_history': payment_history,
            'billing_history': billing_history
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#======================================================================================
# AUTHENTICATION
#======================================================================================

@csrf_exempt
def login_user(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        username = data.get("username")
        password = data.get("password")
        admin = Admin.objects.filter(username=username, password=password).first()
        if admin:
            token = secrets.token_hex(32)
            create_log(
                username="admin",
                role="admin",
                action="LOGIN",
                table="admin",
                record_id=admin.id,
                description="Admin logged into system"
            )
            create_audit_trail(
                username="admin",
                role="admin",
                action="LOGIN",
                description="Admin logged into system",
                request=request
            )
            return JsonResponse({"token": token})
        return JsonResponse({"error": "Invalid login credentials"}, status=401)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def users_login(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST requests allowed"}, status=405)
    try:
        data = json.loads(request.body)
        user = Users.objects.filter(
            username=data.get("username"),
            password=data.get("password")
        ).first()
        if not user:
            return JsonResponse({"error": "Invalid credentials"}, status=401)
        token = secrets.token_hex(16)
        create_log(
            user.username,
            user.role,
            "LOGIN",
            "users",
            user.id,
            f"{user.username} logged into system"
        )
        create_audit_trail(
            username=user.username,
            role=user.role,
            action="LOGIN",
            description=f"{user.username} logged into system",
            request=request
        )
        return JsonResponse({
            "token": token,
            "username": user.username,
            "role": user.role
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# USER MANAGEMENT
#======================================================================================

@csrf_exempt
def new_user(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        fname = data.get("fname")
        phone = data.get("phone")
        metre_num = data.get("metre_num")
        zone = data.get("zone")
        rate = data.get("rate")
        grp = data.get("grp")
        parent = data.get("parent")
        user_name = data.get("username")
        role = data.get("role")
        
        if not all([fname, phone, metre_num, zone, rate]):
            return JsonResponse({"error": "Missing fields"}, status=400)
        
        with transaction.atomic():
            user = read_users.objects.create(
                fname=fname,
                phone=phone,
                metre_num=metre_num,
                zone=zone,
                rate=rate,
                grp=grp,
                parent=parent
            )
            today = date.today()
            reading = readings.objects.create(
                user=user,
                name=fname,
                phone=phone,
                prev_user=0,
                prev_sup=0,
                prev_date=today,
                cur_user=None,
                cur_sup=None,
                cur_date=today,
                units_used=0,
                rate=rate,
                metre_num=metre_num,
                grp=grp,
                parent=parent
            )
            
            # Create initial reading history
            create_reading_history(reading, user_name, role)
            
            create_log(
                username=user_name,
                role=role,
                action="CREATE",
                table="waterusers",
                record_id=user.id,
                description=f"{user_name} created new customer {fname}"
            )
            create_audit_trail(
                username=user_name,
                role=role,
                action="CREATE",
                table_name="waterusers",
                record_id=user.id,
                description=f"{user_name} created new customer {fname}",
                request=request
            )
        
        return JsonResponse({"message": "User registered successfully"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def update_user(request, user_id):
    if request.method != "PUT":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        fname = data.get("fname")
        phone = data.get("phone")
        metre_num = data.get("metre_num")
        zone = data.get("zone")
        rate = data.get("rate")
        grp = data.get("grp")
        parent = data.get("parent")
        user_name = data.get("username", "Unknown")
        role = data.get("role", "Unknown")
        
        with transaction.atomic():
            try:
                user = read_users.objects.get(id=user_id)
            except read_users.DoesNotExist:
                return JsonResponse({"error": "User not found"}, status=404)
            
            old_name = user.fname
            old_phone = user.phone
            
            # Store old user data for audit
            old_data = {
                'fname': user.fname,
                'phone': user.phone,
                'metre_num': user.metre_num,
                'zone': user.zone,
                'rate': user.rate,
                'grp': user.grp,
                'parent': user.parent
            }
            
            user.fname = fname or user.fname
            user.phone = phone or user.phone
            user.metre_num = metre_num or user.metre_num
            user.zone = zone or user.zone
            user.rate = rate or user.rate
            user.grp = grp or user.grp
            user.parent = parent or user.parent
            user.save()
            
            # Update readings and billings
            readings.objects.filter(user_id=user_id).update(
                name=fname,
                phone=phone,
                metre_num=metre_num,
                rate=rate,
                grp=grp,
                parent=parent
            )
            
            Billings.objects.filter(user_id=user_id).update(
                name=fname,
                phone=phone,
                rate=rate,
                sms_name=metre_num,
                grp=grp,
                parent=parent
            )
            
            # Update customer summary
            update_customer_summary(user_id)
            
            create_log(
                username=user_name,
                role=role,
                action="UPDATE",
                table="waterusers",
                record_id=user_id,
                description=f"{role} updated customer {old_name} → {fname}"
            )
            create_audit_trail(
                username=user_name,
                role=role,
                action="UPDATE",
                table_name="waterusers",
                record_id=user_id,
                old_value=json.dumps(old_data),
                new_value=json.dumps({
                    'fname': user.fname,
                    'phone': user.phone,
                    'metre_num': user.metre_num,
                    'zone': user.zone,
                    'rate': user.rate,
                    'grp': user.grp,
                    'parent': user.parent
                }),
                description=f"{role} updated customer {old_name} → {fname}",
                request=request
            )
        
        return JsonResponse({"message": "User updated successfully"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def delete_user(request, user_id):
    if request.method != "DELETE":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body) if request.body else {}
        user_name = data.get("username", "Unknown")
        role = data.get("role", "Unknown")
        
        with transaction.atomic():
            try:
                user = read_users.objects.get(id=user_id)
            except read_users.DoesNotExist:
                return JsonResponse({"error": "User not found"}, status=404)
            
            fname = user.fname
            
            # Get readings for history before deletion
            reading_records = readings.objects.filter(user_id=user_id)
            for r in reading_records:
                create_reading_history(r, user_name, role)
            
            # Delete records
            readings.objects.filter(user_id=user_id).delete()
            Billings.objects.filter(user_id=user_id).delete()
            Billings.objects.filter(name=user.fname).delete()
            
            # Delete customer summary
            CustomerPaymentSummary.objects.filter(user_id=user_id).delete()
            
            user.delete()
            
            create_log(
                username=user_name,
                role=role,
                action="DELETE",
                table="waterusers",
                record_id=user_id,
                description=f"{user_name} deleted customer {fname}, readings, billings"
            )
            create_audit_trail(
                username=user_name,
                role=role,
                action="DELETE",
                table_name="waterusers",
                record_id=user_id,
                description=f"{user_name} deleted customer {fname}",
                request=request
            )
        
        return JsonResponse({"message": "User fully deleted"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# SUBMIT READINGS AND BILLING
#======================================================================================

@csrf_exempt
def submit_new_reading(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        updates = data if isinstance(data, list) else [data]
        
        with transaction.atomic():
            for item in updates:
                user_name = item.get("username", "system")
                role = item.get("role", "system")
                reading = readings.objects.select_for_update().get(
                    user_id=item["user_id"]
                )
                
                # Save current state to history BEFORE updating
                create_reading_history(reading, user_name, role)
                
                cur_user = item.get("cur_user")
                cur_sup = item.get("cur_sup")
                
                if cur_user is not None:
                    cur_user = int(cur_user)
                    reading.units_used = max(0, cur_user - (reading.prev_user or 0))
                    update_reading_field(reading, "cur_user", cur_user, user_name, role)
                
                if cur_sup is not None:
                    update_reading_field(reading, "cur_sup", cur_sup, user_name, role)
                
                reading.mid_user = item.get("mid_user", reading.mid_user)
                reading.mid_sup = item.get("mid_sup", reading.mid_sup)
                reading.save()
                
                # BILLING
                bill_amount = reading.units_used * reading.rate
                if reading.units_used == 0:
                    bill_amount = 300
                
                old_billing = Billings.objects.filter(user_id=reading.user_id).first()
                previous_balance = old_billing.b_cd if old_billing else Decimal("0")
                previous_paid = Decimal("0")
                
                if old_billing:
                    previous_balance = old_billing.bal or Decimal("0")
                    previous_paid = old_billing.paid or Decimal("0")
                
                total_balance = previous_balance + Decimal(str(bill_amount))
                billing = Billings.objects.filter(user_id=reading.user_id).first()
                
                if billing:
                    # Save billing history before update
                    cycle_month = reading.cur_date.strftime("%Y-%m") if reading.cur_date else None
                    create_billing_history(billing, cycle_month, user_name, role)
                    
                    billing.name = reading.name
                    billing.phone = reading.phone
                    billing.units_used = reading.units_used
                    billing.rate = reading.rate
                    billing.bill = bill_amount
                    billing.b_cd = previous_balance
                    billing.bal = total_balance
                    billing.paid = 0
                    billing.status = "Unpaid"
                    billing.prev_user = reading.prev_user
                    billing.cur_user = reading.cur_user
                    billing.sms_name = reading.metre_num
                    billing.grp = reading.grp
                    billing.parent = reading.parent
                    billing.save()
                    
                    # Update customer summary
                    update_customer_summary(reading.user_id)
                else:
                    billing = Billings.objects.create(
                        user_id=reading.user_id,
                        name=reading.name,
                        phone=reading.phone,
                        units_used=reading.units_used,
                        rate=reading.rate,
                        bill=bill_amount,
                        b_cd=previous_balance,
                        bal=total_balance,
                        paid=0,
                        status="Unpaid",
                        prev_user=reading.prev_user,
                        cur_user=reading.cur_user,
                        sms_name=reading.metre_num,
                        grp=reading.grp,
                        parent=reading.parent
                    )
                    cycle_month = reading.cur_date.strftime("%Y-%m") if reading.cur_date else None
                    create_billing_history(billing, cycle_month, user_name, role)
                    
                    # Create initial customer summary
                    update_customer_summary(reading.user_id)
                
                create_audit_trail(
                    username=user_name,
                    role=role,
                    action="UPDATE",
                    table_name="readings",
                    record_id=reading.id,
                    description=f"Reading submitted for {reading.name}",
                    request=request
                )
        
        return JsonResponse({"message": "Saved successfully"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# PAYMENT UPDATES (UPDATED WITH CORRECT MATH)
#======================================================================================

@csrf_exempt
def update_paid(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        
        if isinstance(data, list):
            updated = []
            with transaction.atomic():
                for item in data:
                    billing = Billings.objects.get(id=item.get("id"))
                    old_paid = billing.paid
                    new_paid = Decimal(str(item.get("paid", 0)))
                    amount = new_paid - old_paid
                    
                    if amount > 0:
                        # Get the previous balance (balance before this payment)
                        previous_balance = billing.bal + amount  # Current bal + amount paid
                        
                        # Create payment history with corrected math
                        receipt = create_payment_history(
                            billing=billing,
                            amount=amount,
                            previous_balance=previous_balance,
                            payment_method='BULK',
                            recorded_by=item.get("username", "system"),
                            role=item.get("role", "system")
                        )
                    
                    billing.paid = new_paid
                    total_due = (billing.bill or 0) + (billing.b_cd or 0)
                    billing.bal = total_due - new_paid
                    
                    if new_paid == 0:
                        billing.status = "Unpaid"
                    elif new_paid < billing.bill:
                        billing.status = "Partially Paid"
                    else:
                        billing.status = "Paid"
                    billing.save()
                    
                    # Update customer summary
                    update_customer_summary(billing.user_id)
                    
                    create_log(
                        item.get("username", "system"),
                        item.get("role", "system"),
                        "UPDATE",
                        "billings",
                        billing.id,
                        f"bulk update: {old_paid} → {new_paid}",
                        "paid",
                        old_paid,
                        new_paid
                    )
                    
                    updated.append({
                        "id": billing.id,
                        "paid": billing.paid,
                        "bal": billing.bal,
                        "status": billing.status
                    })
            
            return JsonResponse({
                "message": "Bulk payment updated successfully",
                "updated": updated
            })
        else:
            billing = Billings.objects.get(id=data.get("id"))
            old_paid = billing.paid
            new_paid = Decimal(str(data.get("paid", 0)))
            amount = new_paid - old_paid
            
            if amount > 0:
                # Get the previous balance (balance before this payment)
                previous_balance = billing.bal + amount  # Current bal + amount paid
                
                receipt = create_payment_history(
                    billing=billing,
                    amount=amount,
                    previous_balance=previous_balance,
                    payment_method=data.get("payment_method", 'CASH'),
                    recorded_by=data.get("username", "system"),
                    role=data.get("role", "system"),
                    notes=data.get("notes", None)
                )
            
            billing.paid = new_paid
            total_due = (billing.bill or 0) + (billing.b_cd or 0)
            billing.bal = total_due - new_paid
            
            if new_paid == 0:
                billing.status = "Unpaid"
            elif new_paid < billing.bill:
                billing.status = "Partially Paid"
            else:
                billing.status = "Paid"
            billing.save()
            
            # Update customer summary
            update_customer_summary(billing.user_id)
            
            create_log(
                data.get("username"),
                data.get("role"),
                "UPDATE",
                "billings",
                billing.id,
                f"{data.get('role')} updated payment from {old_paid} to {new_paid}",
                "paid",
                old_paid,
                new_paid
            )
            
            return JsonResponse({
                "message": "Payment updated",
                "id": billing.id,
                "paid": billing.paid,
                "bal": billing.bal,
                "status": billing.status,
                "receipt_number": receipt if amount > 0 else None
            })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# MONTH FINALIZATION
#======================================================================================

@csrf_exempt
def finalize_month(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body) if request.body else {}
        username = data.get("username", "system")
        role = data.get("role", "system")
        
        today = datetime.now()
        last_day = calendar.monthrange(today.year, today.month)[1]
        cycle_end = datetime(today.year, today.month, last_day, 23, 59, 59)
        
        if today < cycle_end:
            return JsonResponse({
                "error": "Cycle not finished yet"
            }, status=400)
        
        next_month = today.month + 1
        next_year = today.year
        if next_month > 12:
            next_month = 1
            next_year += 1
        next_last_day = calendar.monthrange(next_year, next_month)[1]
        next_cycle_date = date(next_year, next_month, next_last_day)
        
        with transaction.atomic():
            # Update billing cycle status
            current_cycle = BillingCycleHistory.objects.filter(
                cycle_month=today.strftime("%Y-%m")
            ).first()
            
            if current_cycle:
                current_cycle.status = 'COMPLETED'
                current_cycle.completed_by = username
                current_cycle.completed_at = timezone.now()
                current_cycle.save()
            
            for r in readings.objects.all():
                # Create history before shift
                create_reading_history(r, username, role)
                
                if r.cur_user is not None:
                    r.prev_user = r.cur_user
                if r.cur_sup is not None:
                    r.prev_sup = r.cur_sup
                
                r.cur_user = None
                r.cur_sup = None
                r.mid_user = 0
                r.mid_sup = 0
                r.prev_date = r.cur_date or r.prev_date
                r.cur_date = next_cycle_date
                r.save()
            
            # Create next billing cycle
            BillingCycleHistory.objects.create(
                cycle_month=next_cycle_date.strftime("%Y-%m"),
                start_date=next_cycle_date,
                end_date=next_cycle_date + timedelta(days=30),
                next_cycle_date=next_cycle_date + timedelta(days=60),
                status='PENDING',
                started_by=username
            )
            
            create_audit_trail(
                username=username,
                role=role,
                action="SYSTEM",
                table_name="readings",
                description=f"Month finalized for {today.strftime('%Y-%m')}",
                request=request
            )
        
        return JsonResponse({
            "message": "Cycle shifted successfully",
            "next_cycle_date": str(next_cycle_date)
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# EMPLOYEE MANAGEMENT
#======================================================================================

@api_view(['POST'])
def register_user(request):
    user = Users.objects.create(
        username=request.data.get('username'),
        password=request.data.get('password'),
        role=request.data.get('role')
    )
    create_log(
        "Admin",
        "admin",
        "CREATE",
        "users",
        user.id,
        f"Admin created employee {user.username}"
    )
    create_audit_trail(
        username="Admin",
        role="admin",
        action="CREATE",
        table_name="users",
        record_id=user.id,
        description=f"Admin created employee {user.username}",
        request=request
    )
    return Response({"message": "User registered successfully"})

@api_view(['GET'])
def list_employees(request):
    return Response(list(Users.objects.all().values('id', 'username', 'role')))

@csrf_exempt
def delete_employee(request, emp_id):
    if request.method != "DELETE":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        emp = Users.objects.get(id=emp_id)
        emp.delete()
        return JsonResponse({"message": "Employee deleted"})
    except Users.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

@csrf_exempt
def update_employee(request, emp_id):
    if request.method != "PUT":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body)
        emp = Users.objects.get(id=emp_id)
        emp.username = data.get("username", emp.username)
        emp.role = data.get("role", emp.role)
        emp.save()
        return JsonResponse({"message": "Employee updated"})
    except Users.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

#======================================================================================
# ANALYTICS
#======================================================================================

def total_bill(request):
    total = Billings.objects.aggregate(total_bill=Sum('bill'))['total_bill'] or 0
    return JsonResponse({"total_bill": round(float(total), 2)})

def total_bal(request):
    total = Billings.objects.aggregate(total_bal=Sum('b_cd'))['total_bal'] or 0
    return JsonResponse({"total_bal": round(float(total), 2)})

def total_paid(request):
    total = Billings.objects.aggregate(total_paid=Sum('paid'))['total_paid'] or 0
    return JsonResponse({"total_paid": round(float(total), 2)})

def total_units(request):
    total = Billings.objects.aggregate(total_units=Sum('units_used'))['total_units'] or 0
    return JsonResponse({"total_units": round(float(total), 2)})

def total_cust(request):
    total = read_users.objects.aggregate(total_cust=Count('id'))['total_cust'] or 0
    return JsonResponse({"total_cust": total})

def avg_units(request):
    avg = Billings.objects.aggregate(avg_units=Avg('units_used'))['avg_units'] or 0
    return JsonResponse({"avg_units": round(float(avg), 2)})

#======================================================================================
# EXCEL UPLOAD/DOWNLOAD
#======================================================================================

def download_readings_template(request):
    template_path = os.path.join(
        settings.BASE_DIR,
        "templates",
        "readings_template.xlsx"
    )
    wb = load_workbook(template_path)
    ws = wb.active
    
    readings_data = readings.objects.all().values(
        "user_id",
        "name",
        "phone",
        "metre_num",
        "prev_user",
        "prev_sup"
    )
    
    row = 2
    for r in readings_data:
        ws[f"A{row}"] = r["user_id"]
        ws[f"B{row}"] = r["name"]
        ws[f"C{row}"] = r["phone"]
        ws[f"D{row}"] = r["metre_num"]
        ws[f"E{row}"] = r["prev_user"]
        ws[f"F{row}"] = r["prev_sup"]
        ws[f"G{row}"] = None
        ws[f"H{row}"] = None
        ws[f"I{row}"] = None
        ws[f"J{row}"] = None
        row += 1
    
    while row <= ws.max_row:
        for col in ["A","B","C","D","E","F","G","H","I","J"]:
            ws[f"{col}{row}"] = None
        row += 1
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="readings_template.xlsx"'
    wb.save(response)
    return response

@csrf_exempt
def upload_readings_excel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        file = request.FILES.get("file")
        if not file:
            return JsonResponse({"error": "No file uploaded"}, status=400)
        
        df = pd.read_excel(file)
        if "user_id" not in df.columns:
            return JsonResponse(
                {"error": "Excel must contain user_id column"},
                status=400
            )
        
        processed = 0
        skipped = 0
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    user_id = row.get("user_id")
                    if pd.isna(user_id):
                        skipped += 1
                        continue
                    
                    reading = readings.objects.select_for_update().get(
                        user_id=int(user_id)
                    )
                    
                    # Create history before update
                    create_reading_history(reading, "excel_upload", "system")
                    
                    cur_user = None if pd.isna(row.get("cur_user")) else int(row.get("cur_user"))
                    cur_sup = None if pd.isna(row.get("cur_sup")) else int(row.get("cur_sup"))
                    mid_user = None if pd.isna(row.get("mid_user")) else int(row.get("mid_user"))
                    mid_sup = None if pd.isna(row.get("mid_sup")) else int(row.get("mid_sup"))
                    
                    if cur_user is None and cur_sup is None and mid_user is None and mid_sup is None:
                        skipped += 1
                        continue
                    
                    if cur_user is not None:
                        reading.cur_user = cur_user
                        reading.units_used = max(0, cur_user - (reading.prev_user or 0))
                    if cur_sup is not None:
                        reading.cur_sup = cur_sup
                    if mid_user is not None:
                        reading.mid_user = mid_user
                    if mid_sup is not None:
                        reading.mid_sup = mid_sup
                    reading.save()
                    
                    # BILLING
                    bill_amount = reading.units_used * reading.rate
                    if reading.units_used <= 2:
                        bill_amount = 300
                    
                    old_billing = Billings.objects.filter(user_id=reading.user_id).first()
                    previous_balance = old_billing.bal if old_billing else Decimal("0")
                    total_balance = previous_balance + Decimal(str(bill_amount))
                    
                    billing = Billings.objects.filter(user_id=reading.user_id).first()
                    if billing:
                        cycle_month = reading.cur_date.strftime("%Y-%m") if reading.cur_date else None
                        create_billing_history(billing, cycle_month, "excel_upload", "system")
                        
                        billing.name = reading.name
                        billing.phone = reading.phone
                        billing.units_used = reading.units_used
                        billing.rate = reading.rate
                        billing.bill = bill_amount
                        billing.b_cd = previous_balance
                        billing.bal = total_balance
                        billing.paid = 0
                        billing.status = "Unpaid"
                        billing.prev_user = reading.prev_user
                        billing.cur_user = reading.cur_user
                        billing.sms_name = reading.metre_num
                        billing.grp = reading.grp
                        billing.parent = reading.parent
                        billing.save()
                    else:
                        billing = Billings.objects.create(
                            user_id=reading.user_id,
                            name=reading.name,
                            phone=reading.phone,
                            units_used=reading.units_used,
                            rate=reading.rate,
                            bill=bill_amount,
                            b_cd=previous_balance,
                            bal=total_balance,
                            paid=0,
                            status="Unpaid",
                            prev_user=reading.prev_user,
                            cur_user=reading.cur_user,
                            sms_name=reading.metre_num,
                            grp=reading.grp,
                            parent=reading.parent
                        )
                        cycle_month = reading.cur_date.strftime("%Y-%m") if reading.cur_date else None
                        create_billing_history(billing, cycle_month, "excel_upload", "system")
                    
                    update_customer_summary(reading.user_id)
                    processed += 1
                except Exception as row_error:
                    print(f"Row {index}: {row_error}")
                    skipped += 1
        
        return JsonResponse({
            "message": "Excel uploaded successfully",
            "processed_rows": processed,
            "skipped_rows": skipped
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def download_billings_template(request):
    data = Billings.objects.all().values(
        "id", "name", "phone", "bill", "paid"
    )
    df = pd.DataFrame(list(data))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=billings_template.xlsx'
    df.to_excel(response, index=False)
    return response

@csrf_exempt
def upload_billings_excel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        file = request.FILES["file"]
        df = pd.read_excel(file)
        updated = []
        
        with transaction.atomic():
            for _, row in df.iterrows():
                billing_id = row.get("id")
                paid = row.get("paid")
                if pd.isna(billing_id) or pd.isna(paid):
                    continue
                
                billing = Billings.objects.get(id=int(billing_id))
                old_paid = billing.paid
                new_paid = Decimal(str(paid))
                amount = new_paid - old_paid
                
                if amount > 0:
                    # Get the previous balance (balance before this payment)
                    previous_balance = billing.bal + amount  # Current bal + amount paid
                    
                    create_payment_history(
                        billing=billing,
                        amount=amount,
                        previous_balance=previous_balance,
                        payment_method='EXCEL',
                        recorded_by="excel_upload",
                        role="system"
                    )
                
                billing.paid = new_paid
                total_due = (billing.bill or 0) + (billing.b_cd or 0)
                billing.bal = total_due - new_paid
                
                if new_paid == 0:
                    billing.status = "Unpaid"
                elif new_paid < billing.bill:
                    billing.status = "Partially Paid"
                else:
                    billing.status = "Paid"
                billing.save()
                
                update_customer_summary(billing.user_id)
                
                create_log(
                    "excel_upload",
                    "system",
                    "UPDATE",
                    "billings",
                    billing.id,
                    f"Excel update: {old_paid} → {new_paid}",
                    "paid",
                    old_paid,
                    new_paid
                )
                updated.append(billing.id)
        
        return JsonResponse({
            "message": "Excel uploaded successfully",
            "updated_count": len(updated)
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def download_users_excel(request):
    users = read_users.objects.all().values(
        "id", "fname", "phone", "metre_num",
        "zone", "rate", "grp", "parent", "created_on"
    )
    df = pd.DataFrame(list(users))
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="water_users.xlsx"'
    df.to_excel(response, index=False)
    return response

#======================================================================================
# OTHER UTILITY ENDPOINTS
#======================================================================================

def billing_timer(request):
    today = datetime.now()
    next_month = today.month + 1
    next_year = today.year
    if next_month > 12:
        next_month = 1
        next_year += 1
    end_date = datetime(next_year, next_month, calendar.monthrange(next_year, next_month)[1], 23, 59, 59)
    
    diff = end_date - today
    return JsonResponse({
        "days": diff.days,
        "hours": diff.seconds // 3600,
        "minutes": (diff.seconds % 3600) // 60,
        "seconds": diff.seconds % 60
    })

@csrf_exempt
def reset_mid_month_readings(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    try:
        data = json.loads(request.body) if request.body else {}
        username = data.get("username", "system")
        role = data.get("role", "system")
        
        with transaction.atomic():
            readings_qs = readings.objects.all()
            for r in readings_qs:
                old_mid_user = r.mid_user
                old_mid_sup = r.mid_sup
                
                # Create history before reset
                create_reading_history(r, username, role)
                
                r.mid_user = 0
                r.mid_sup = 0
                r.save()
                
                create_log(
                    username=username,
                    role=role,
                    action="UPDATE",
                    table="readings",
                    record_id=r.id,
                    field_changed="mid_month_reset",
                    old_val=f"user:{old_mid_user}, sup:{old_mid_sup}",
                    new_val="user:0, sup:0",
                    description=f"Mid-month readings reset for {r.name}"
                )
        
        return JsonResponse({
            "message": "Mid-month readings reset successfully"
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def restore_readings(request):
    global LAST_STATE_SNAPSHOT
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)
    if not LAST_STATE_SNAPSHOT:
        return JsonResponse({"error": "No snapshot available"}, status=400)
    
    with transaction.atomic():
        readings.objects.all().delete()
        for r in LAST_STATE_SNAPSHOT:
            r.pop("id", None)
            readings.objects.create(**r)
    
    return JsonResponse({"message": "System restored successfully"})

@csrf_exempt
def update_all_users(request):
    if request.method != "PUT":
        return JsonResponse({"error": "PUT request required"}, status=405)
    try:
        customers = json.loads(request.body)
        with transaction.atomic():
            for customer in customers:
                user = read_users.objects.get(id=customer["id"])
                
                old_data = {
                    'fname': user.fname,
                    'phone': user.phone,
                    'metre_num': user.metre_num,
                    'rate': user.rate,
                    'grp': user.grp,
                    'parent': user.parent
                }
                
                user.fname = customer["fname"]
                user.phone = customer["phone"]
                user.metre_num = customer["metre_num"]
                user.rate = customer["rate"]
                user.grp = customer["grp"]
                user.parent = customer["parent"]
                user.save()
                
                readings.objects.filter(user_id=user.id).update(
                    name=user.fname,
                    phone=user.phone,
                    metre_num=user.metre_num,
                    rate=user.rate,
                    grp=user.grp,
                    parent=user.parent
                )
                
                Billings.objects.filter(user_id=user.id).update(
                    name=user.fname,
                    phone=user.phone,
                    rate=user.rate,
                    sms_name=user.metre_num,
                    grp=user.grp,
                    parent=user.parent
                )
                
                update_customer_summary(user.id)
        
        return JsonResponse({
            "success": True,
            "message": "All users updated successfully."
        })
    except read_users.DoesNotExist:
        return JsonResponse({"error": "One or more users do not exist."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

@csrf_exempt
def update_all_bill_phones(request):
    if request.method != "PUT":
        return JsonResponse({"error": "PUT request required"}, status=405)
    try:
        data = json.loads(request.body)
        phone = data.get("phone")
        if not phone:
            return JsonResponse({"error": "Phone number is required"}, status=400)
        
        updated = Billings.objects.update(phone=phone)
        
        # Also update customer summaries
        CustomerPaymentSummary.objects.update(phone=phone)
        
        return JsonResponse({
            "success": True,
            "updated": updated
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

#======================================================================================
# SMS FUNCTIONS
#======================================================================================

API_URL = "https://quicksms.advantasms.com/api/services/sendbulk"
PARTNER_ID = "16256"
API_KEY = "bc1bc562ccb7c72732e7fa0add447129"
SHORTCODE = "AdvantaSMS"

def send_bulk_sms(customers):
    payload = {
        "count": len(customers),
        "smslist": []
    }
    for i, customer in enumerate(customers):
        payload["smslist"].append({
            "partnerID": PARTNER_ID,
            "apikey": API_KEY,
            "pass_type": "plain",
            "clientsmsid": i + 1,
            "mobile": customer["phone"],
            "message": customer["message"],
            "shortcode": SHORTCODE
        })
    headers = {
        "Content-Type": "application/json"
    }
    response = requests.post(API_URL, json=payload, headers=headers)
    return response.json()

@csrf_exempt
def send_sms_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            customers = data.get("customers", [])
            if not customers:
                return JsonResponse({"error": "No customers selected"}, status=400)
            
            result = send_bulk_sms(customers)
            print(result)
            
            # Log SMS sending
            create_audit_trail(
                username=data.get("username", "system"),
                role=data.get("role", "system"),
                action="EXPORT",
                table_name="sms",
                description=f"SMS sent to {len(customers)} customers",
                request=request
            )
            
            return JsonResponse(result, safe=False)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

def process_reading_update(
    user_id,
    new_cur_user=None,
    new_cur_sup=None,
    mid_user=None,
    mid_sup=None,
    username="system",
    role="system"
):
    try:
        reading = readings.objects.get(user_id=user_id)
    except readings.DoesNotExist:
        create_log(username, role, "ERROR", "readings", user_id,
                   "Reading record not found")
        return
    
    # Create history before any changes
    create_reading_history(reading, username, role)
    
    prev_user = reading.prev_user or 0
    prev_sup = reading.prev_sup or 0
    
    if mid_user is not None:
        create_log(username, role, "UPDATE", "readings", reading.id,
                   f"mid_user {reading.mid_user} → {mid_user}",
                   "mid_user", reading.mid_user, mid_user)
        update_reading_field(reading, "mid_user", mid_user, username, role)
    
    if mid_sup is not None:
        create_log(username, role, "UPDATE", "readings", reading.id,
                   f"mid_sup {reading.mid_sup} → {mid_sup}",
                   "mid_sup", reading.mid_sup, mid_sup)
        update_reading_field(reading, "mid_sup", mid_sup, username, role)
    
    if new_cur_user is None and new_cur_sup is None:
        reading.save()
        return
    
    units_used = reading.units_used or 0
    if new_cur_user is not None:
        try:
            units_used = int(new_cur_user) - int(prev_user)
        except Exception:
            create_log(username, role, "ERROR", "readings", reading.id,
                       "Invalid numeric reading input")
            return
        if units_used < 0:
            units_used = 0
    
    if new_cur_user is not None:
        create_log(username, role, "UPDATE", "readings", reading.id,
                   f"user reading {prev_user} → {new_cur_user}",
                   "cur_user", prev_user, new_cur_user)
    
    if new_cur_sup is not None:
        create_log(username, role, "UPDATE", "readings", reading.id,
                   f"sup reading {prev_sup} → {new_cur_sup}",
                   "cur_sup", prev_sup, new_cur_sup)
    
    if new_cur_user is not None:
        update_reading_field(reading, "cur_user", new_cur_user, username, role)
    if new_cur_sup is not None:
        update_reading_field(reading, "cur_sup", new_cur_sup, username, role)
    
    reading.units_used = units_used
    reading.save()
    
    # BILLING
    if new_cur_user is not None:
        rate = reading.rate or 0
        bill_amount = units_used * rate
        if units_used <= 2:
            bill_amount = 300
        
        old_billing = Billings.objects.filter(user_id=user_id).first()
        previous_balance = old_billing.bal if old_billing else Decimal("0")
        
        billing, created = Billings.objects.get_or_create(
            user_id=user_id,
            billed_on=date.today(),
            defaults={
                "name": reading.name,
                "phone": reading.phone,
                "units_used": units_used,
                "rate": rate,
                "bill": bill_amount,
                "paid": 0,
                "bal": bill_amount,
                "status": "Unpaid",
                "b_cd": previous_balance,
                "prev_user": reading.prev_user,
                "cur_user": reading.cur_user,
                "sms_name": reading.metre_num,
                "grp": reading.grp,
                "parent": reading.parent
            }
        )
        
        if not created:
            # Save billing history before update
            cycle_month = reading.cur_date.strftime("%Y-%m") if reading.cur_date else None
            create_billing_history(billing, cycle_month, username, role)
            
            billing.units_used = units_used
            billing.bill = bill_amount
            billing.paid = 0
            billing.b_cd = previous_balance
            billing.bal = previous_balance + Decimal(str(bill_amount))
            
            if billing.paid == 0:
                billing.status = "Unpaid"
            elif billing.paid < bill_amount:
                billing.status = "Partially Paid"
            else:
                billing.status = "Paid"
            billing.save()
        
        update_customer_summary(user_id)

# ============================================================
# PAYMENT HISTORY FETCH ENDPOINTS
# ============================================================

@api_view(['GET'])
def get_all_payment_history(request):
    """
    Fetch all payment history records with optional filters.
    Returns complete payment history with customer details.
    """
    try:
        # Get filter parameters
        user_id = request.GET.get('user_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        payment_method = request.GET.get('payment_method')
        status = request.GET.get('status')
        search = request.GET.get('search')  # Search by name or phone
        
        # Start with all payment history
        payment_history_qs = PaymentHistory.objects.all()
        
        # Apply filters
        if user_id:
            payment_history_qs = payment_history_qs.filter(user_id=user_id)
        
        if start_date:
            payment_history_qs = payment_history_qs.filter(payment_date__gte=start_date)
        
        if end_date:
            payment_history_qs = payment_history_qs.filter(payment_date__lte=end_date)
        
        if payment_method:
            payment_history_qs = payment_history_qs.filter(payment_method=payment_method)
        
        if status:
            payment_history_qs = payment_history_qs.filter(status=status)
        
        if search:
            payment_history_qs = payment_history_qs.filter(
                Q(name__icontains=search) | Q(phone__icontains=search)
            )
        
        # Order by most recent first
        payment_history_qs = payment_history_qs.order_by('-timestamp')
        
        # Prepare data for frontend
        data = []
        for p in payment_history_qs:
            data.append({
                'id': p.id,
                'billing_id': p.billing_id,
                'user_id': p.user_id,
                'name': p.name,
                'phone': p.phone,
                'grp': p.grp,
                'parent': p.parent,
                'amount_paid': float(p.amount_paid),
                'previous_balance': float(p.previous_balance),
                'current_balance': float(p.current_balance),
                'bill_amount': float(p.bill_amount),
                'payment_method': p.payment_method,
                'payment_method_display': dict(PaymentHistory.PAYMENT_METHODS).get(p.payment_method, p.payment_method),
                'status': p.status,
                'status_display': dict(PaymentHistory.PAYMENT_STATUS).get(p.status, p.status),
                'receipt_number': p.receipt_number,
                'notes': p.notes,
                'payment_date': p.payment_date.strftime('%Y-%m-%d') if p.payment_date else None,
                'recorded_by': p.recorded_by,
                'role': p.role,
                'timestamp': p.timestamp.strftime('%Y-%m-%d %H:%M:%S') if p.timestamp else None
            })
        
        # Get summary statistics
        total_payments = payment_history_qs.count()
        total_amount = payment_history_qs.aggregate(
            total=Sum('amount_paid')
        )['total'] or 0
        
        return Response({
            'success': True,
            'data': data,
            'summary': {
                'total_payments': total_payments,
                'total_amount': float(total_amount),
                'filters_applied': {
                    'user_id': user_id,
                    'start_date': start_date,
                    'end_date': end_date,
                    'payment_method': payment_method,
                    'status': status,
                    'search': search
                }
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_payment_history_by_user(request, user_id):
    """
    Fetch payment history for a specific user.
    """
    try:
        payment_history = PaymentHistory.objects.filter(
            user_id=user_id
        ).order_by('-timestamp')
        
        if not payment_history.exists():
            return Response({
                'success': True,
                'data': [],
                'message': 'No payment history found for this user'
            })
        
        data = []
        for p in payment_history:
            data.append({
                'id': p.id,
                'billing_id': p.billing_id,
                'amount_paid': float(p.amount_paid),
                'previous_balance': float(p.previous_balance),
                'current_balance': float(p.current_balance),
                'bill_amount': float(p.bill_amount),
                'payment_method': p.payment_method,
                'payment_method_display': dict(PaymentHistory.PAYMENT_METHODS).get(p.payment_method, p.payment_method),
                'status': p.status,
                'status_display': dict(PaymentHistory.PAYMENT_STATUS).get(p.status, p.status),
                'receipt_number': p.receipt_number,
                'payment_date': p.payment_date.strftime('%Y-%m-%d') if p.payment_date else None,
                'recorded_by': p.recorded_by,
                'timestamp': p.timestamp.strftime('%Y-%m-%d %H:%M:%S') if p.timestamp else None
            })
        
        # Get user summary
        user_summary = CustomerPaymentSummary.objects.filter(user_id=user_id).first()
        
        return Response({
            'success': True,
            'user_id': user_id,
            'payment_history': data,
            'summary': {
                'total_paid': float(user_summary.total_paid) if user_summary else 0,
                'current_balance': float(user_summary.current_balance) if user_summary else 0,
                'payment_count': len(data),
                'last_payment': data[0] if data else None
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_payment_summary(request):
    """
    Get summary statistics of all payments.
    """
    try:
        # Get filter parameters
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        payment_history_qs = PaymentHistory.objects.all()
        
        if start_date:
            payment_history_qs = payment_history_qs.filter(payment_date__gte=start_date)
        if end_date:
            payment_history_qs = payment_history_qs.filter(payment_date__lte=end_date)
        
        # Aggregations
        total_payments = payment_history_qs.count()
        total_amount = payment_history_qs.aggregate(
            total=Sum('amount_paid')
        )['total'] or 0
        
        # Payment method breakdown
        method_breakdown = payment_history_qs.values('payment_method').annotate(
            count=Count('id'),
            total=Sum('amount_paid')
        ).order_by('-total')
        
        # Daily payment trend
        daily_trend = payment_history_qs.values('payment_date').annotate(
            count=Count('id'),
            total=Sum('amount_paid')
        ).order_by('-payment_date')[:30]  # Last 30 days
        
        # Prepare method breakdown with display names
        method_data = []
        for method in method_breakdown:
            method_data.append({
                'method': method['payment_method'],
                'method_display': dict(PaymentHistory.PAYMENT_METHODS).get(method['payment_method'], method['payment_method']),
                'count': method['count'],
                'total': float(method['total'])
            })
        
        return Response({
            'success': True,
            'summary': {
                'total_payments': total_payments,
                'total_amount': float(total_amount),
                'average_payment': float(total_amount / total_payments) if total_payments > 0 else 0
            },
            'method_breakdown': method_data,
            'daily_trend': list(daily_trend)
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_payment_receipt(request, receipt_number):
    """
    Get payment details by receipt number.
    """
    try:
        payment = PaymentHistory.objects.filter(
            receipt_number=receipt_number
        ).first()
        
        if not payment:
            return Response({
                'success': False,
                'error': 'Payment receipt not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get associated billing information
        billing = Billings.objects.filter(id=payment.billing_id).first()
        
        data = {
            'id': payment.id,
            'billing_id': payment.billing_id,
            'user_id': payment.user_id,
            'name': payment.name,
            'phone': payment.phone,
            'grp': payment.grp,
            'parent': payment.parent,
            'amount_paid': float(payment.amount_paid),
            'previous_balance': float(payment.previous_balance),
            'current_balance': float(payment.current_balance),
            'bill_amount': float(payment.bill_amount),
            'payment_method': payment.payment_method,
            'payment_method_display': dict(PaymentHistory.PAYMENT_METHODS).get(payment.payment_method, payment.payment_method),
            'status': payment.status,
            'status_display': dict(PaymentHistory.PAYMENT_STATUS).get(payment.status, payment.status),
            'receipt_number': payment.receipt_number,
            'notes': payment.notes,
            'payment_date': payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else None,
            'recorded_by': payment.recorded_by,
            'role': payment.role,
            'timestamp': payment.timestamp.strftime('%Y-%m-%d %H:%M:%S') if payment.timestamp else None,
            'billing_details': {
                'units_used': billing.units_used if billing else None,
                'rate': billing.rate if billing else None,
                'bill_amount': billing.bill if billing else None,
                'status': billing.status if billing else None
            } if billing else None
        }
        
        return Response({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================
# SIMPLE JSON RESPONSE VERSION (if you prefer not to use DRF)
# ============================================================

def get_payment_history_json(request):
    """
    Simple JSON response version of payment history.
    This uses regular Django JsonResponse instead of DRF.
    """
    try:
        # Get all payment history
        payments = PaymentHistory.objects.all().order_by('-timestamp')
        
        # Prepare data
        data = []
        for p in payments:
            data.append({
                'id': p.id,
                'billing_id': p.billing_id,
                'user_id': p.user_id,
                'name': p.name,
                'phone': p.phone,
                'grp': p.grp,
                'parent': p.parent,
                'amount_paid': float(p.amount_paid),
                'previous_balance': float(p.previous_balance),
                'current_balance': float(p.current_balance),
                'bill_amount': float(p.bill_amount),
                'payment_method': p.payment_method,
                'payment_method_display': dict(PaymentHistory.PAYMENT_METHODS).get(p.payment_method, p.payment_method),
                'status': p.status,
                'status_display': dict(PaymentHistory.PAYMENT_STATUS).get(p.status, p.status),
                'receipt_number': p.receipt_number,
                'notes': p.notes,
                'payment_date': p.payment_date.strftime('%Y-%m-%d') if p.payment_date else None,
                'recorded_by': p.recorded_by,
                'role': p.role,
                'timestamp': p.timestamp.strftime('%Y-%m-%d %H:%M:%S') if p.timestamp else None
            })
        
        # Get summary
        total_amount = PaymentHistory.objects.aggregate(
            total=Sum('amount_paid')
        )['total'] or 0
        
        return JsonResponse({
            'success': True,
            'data': data,
            'summary': {
                'total_payments': len(data),
                'total_amount': float(total_amount)
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================
# PAYMENT RECEIPT DOWNLOAD
# ============================================================

def download_payment_receipt(request, receipt_number):
    """
    Download payment receipt as PDF.
    """
    try:
        # Get payment record
        payment = PaymentHistory.objects.filter(
            receipt_number=receipt_number
        ).first()
        
        if not payment:
            return JsonResponse({
                'success': False,
                'error': 'Payment receipt not found'
            }, status=404)
        
        # Get billing info
        billing = Billings.objects.filter(id=payment.billing_id).first()
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            alignment=1,  # Center
            spaceAfter=30
        )
        
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=12
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=6
        )
        
        # Build PDF content
        elements = []
        
        # Header
        elements.append(Paragraph("WATER BILLING SYSTEM", title_style))
        elements.append(Paragraph("Payment Receipt", heading_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Receipt details
        receipt_data = [
            ['Receipt Number:', payment.receipt_number],
            ['Date:', payment.payment_date.strftime('%Y-%m-%d %H:%M:%S') if payment.payment_date else 'N/A'],
            ['Payment Method:', payment.payment_method],
            ['Status:', payment.status],
            ['Recorded By:', payment.recorded_by],
        ]
        
        receipt_table = Table(receipt_data, colWidths=[2*inch, 4*inch])
        receipt_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ]))
        elements.append(receipt_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Customer details
        elements.append(Paragraph("Customer Details", heading_style))
        customer_data = [
            ['Name:', payment.name],
            ['Phone:', payment.phone],
            ['Group:', payment.grp or 'N/A'],
            ['Parent:', payment.parent or 'N/A'],
        ]
        
        customer_table = Table(customer_data, colWidths=[2*inch, 4*inch])
        customer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ]))
        elements.append(customer_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Payment details
        elements.append(Paragraph("Payment Details", heading_style))
        payment_data = [
            ['Previous Balance:', f"KES {float(payment.previous_balance):,.2f}"],
            ['Bill Amount:', f"KES {float(payment.bill_amount):,.2f}"],
            ['Amount Paid:', f"KES {float(payment.amount_paid):,.2f}"],
            ['Current Balance:', f"KES {float(payment.current_balance):,.2f}"],
        ]
        
        # Highlight current balance
        payment_table = Table(payment_data, colWidths=[2*inch, 4*inch])
        payment_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (1, 3), (1, 3), colors.red if payment.current_balance > 0 else colors.green),
            ('FONTNAME', (1, 3), (1, 3), 'Helvetica-Bold'),
        ]))
        elements.append(payment_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Billing details if available
        if billing:
            elements.append(Paragraph("Billing Details", heading_style))
            billing_data = [
                ['Units Used:', f"{billing.units_used or 0} m³"],
                ['Rate:', f"KES {float(billing.rate or 0):,.2f} per m³"],
                ['Bill Status:', billing.status or 'N/A'],
            ]
            
            billing_table = Table(billing_data, colWidths=[2*inch, 4*inch])
            billing_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ]))
            elements.append(billing_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Notes if any
        if payment.notes:
            elements.append(Paragraph("Notes:", heading_style))
            elements.append(Paragraph(payment.notes, normal_style))
            elements.append(Spacer(1, 0.2*inch))
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1,
        )
        elements.append(Paragraph("Thank you for your payment!", footer_style))
        elements.append(Paragraph("Kamengo agencies", footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{receipt_number}.pdf"'
        
        # Log the download
        create_audit_trail(
            username=request.GET.get('username', 'system'),
            role=request.GET.get('role', 'system'),
            action="DOWNLOAD",
            table_name="payment_history",
            record_id=payment.id,
            description=f"Payment receipt downloaded: {receipt_number}",
            request=request
        )
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)